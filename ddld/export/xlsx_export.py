"""
Export the full run to Excel (.xlsx): a Claims sheet with color-filled verdict
cells and a Transcript sheet. Requires openpyxl.
"""
from __future__ import annotations

from ..types import CheckedClaim, Utterance, Verdict


def export_xlsx(
    path: str,
    transcript: list[Utterance],
    checked: list[CheckedClaim],
    disclaimer: str,
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Claims sheet ----
    ws = wb.active
    ws.title = "Claims"
    ws["A1"] = "DISCLAIMER: " + disclaimer
    ws["A1"].font = Font(italic=True, size=9)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 42

    headers = ["Time (s)", "Speaker", "Claim", "Verdict", "Confidence", "Reasoning", "Sources", "Checked at"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)

    for c in checked:
        r = c.as_row()
        ws.append([
            r["timestamp_s"], r["speaker"], r["claim"], r["verdict_label"],
            r["confidence"] if c.verdict is not Verdict.PENDING else None,
            r["reasoning"], r["sources"], r["checked_at"],
        ])
        vcell = ws.cell(row=ws.max_row, column=4)
        vcell.fill = PatternFill("solid", fgColor=c.verdict.hex)
        vcell.font = Font(bold=True)

    widths = [10, 16, 60, 22, 11, 55, 50, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A3"

    # ---- Transcript sheet ----
    ts = wb.create_sheet("Transcript")
    ts.append(["Start (s)", "End (s)", "Speaker", "Text"])
    for cell in ts[1]:
        cell.font = Font(bold=True)
    for u in transcript:
        ts.append([round(u.start, 1), round(u.end, 1), u.speaker or "", u.text])
    for col, w in zip("ABCD", [10, 10, 16, 100]):
        ts.column_dimensions[col].width = w
    ts.freeze_panes = "A2"

    wb.save(path)
    return path
